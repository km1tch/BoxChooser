"""
Excel import/export functionality for box prices.
Handles all Excel-related operations using openpyxl.
"""

import json
import os
import re
import tempfile
from datetime import datetime
from typing import Dict, List, Any, Optional

from fastapi import HTTPException, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from backend.lib.box_library import BoxLibrary


def export_prices_to_excel(store_id: str, store_data: dict) -> FileResponse:
    """Export store prices to Excel file"""
    
    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = f'Store {store_id} Prices'
    
    # Define headers for itemized pricing
    headers = ["Model", "Dimensions", "Type", "Box Price", 
              "Standard Materials", "Standard Services", "Standard Total",
              "Fragile Materials", "Fragile Services", "Fragile Total",
              "Custom Materials", "Custom Services", "Custom Total", "Location"]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write data
    row_idx = 2
    for box in store_data["boxes"]:
        box_model = box.get("model", f"Unknown-{len(box['dimensions'])}")
        dimensions = "x".join(map(str, box["dimensions"]))
        
        # Start with common columns
        ws.cell(row=row_idx, column=1, value=box_model)
        ws.cell(row=row_idx, column=2, value=dimensions)
        ws.cell(row=row_idx, column=3, value=box["type"])
        
        # Itemized pricing
        ip = box.get("itemized-prices", {})
        box_price = ip.get("box-price", 0)
        
        ws.cell(row=row_idx, column=4, value=box_price)
        ws.cell(row=row_idx, column=5, value=ip.get("standard-materials", 0))
        ws.cell(row=row_idx, column=6, value=ip.get("standard-services", 0))
        ws.cell(row=row_idx, column=7, value=box_price + ip.get("standard-materials", 0) + ip.get("standard-services", 0))
        ws.cell(row=row_idx, column=8, value=ip.get("fragile-materials", 0))
        ws.cell(row=row_idx, column=9, value=ip.get("fragile-services", 0))
        ws.cell(row=row_idx, column=10, value=box_price + ip.get("fragile-materials", 0) + ip.get("fragile-services", 0))
        ws.cell(row=row_idx, column=11, value=ip.get("custom-materials", 0))
        ws.cell(row=row_idx, column=12, value=ip.get("custom-services", 0))
        ws.cell(row=row_idx, column=13, value=box_price + ip.get("custom-materials", 0) + ip.get("custom-services", 0))
        
        # Location
        location = box.get("location", "")
        if isinstance(location, dict):
            location = location.get("label", "")
        ws.cell(row=row_idx, column=14, value=location)
        
        row_idx += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    filename = f"store{store_id}_prices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(temp_file.name)
    
    return FileResponse(
        temp_file.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


async def import_prices_from_excel(
    store_id: str, 
    file: UploadFile, 
    current_data: dict,
    save_yaml_func
) -> dict:
    """Import prices from Excel file"""
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    # Check file size (512KB limit for production)
    MAX_FILE_SIZE = 512 * 1024  # 512KB
    
    # Stream file to temp location to avoid loading into memory
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        # Stream file content to temp file in chunks
        total_size = 0
        chunk_size = 8192  # 8KB chunks
        
        while True:
            chunk = await file.read(chunk_size)
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > MAX_FILE_SIZE:
                temp_file.close()
                os.unlink(temp_file.name)
                raise HTTPException(status_code=400, detail="File size exceeds 512KB limit")
            temp_file.write(chunk)
        
        temp_file.close()
        await file.seek(0)  # Reset file pointer
        
        # Read Excel file with minimal memory usage
        wb = load_workbook(temp_file.name, read_only=True, keep_vba=False, data_only=True)
        ws = wb.active
        
        # Load current store data
        
        # Get headers from first row
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value] = col_idx
        
        # Create a mapping of models to boxes for faster lookup
        box_map = {}
        for i, box in enumerate(current_data["boxes"]):
            model = box.get("model", f"Unknown-{len(box['dimensions'])}")
            box_map[model] = i
        
        # Process imported data
        updated_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # Get model from row
                model_col = headers.get("Model")
                if not model_col or row_idx > ws.max_row:
                    break
                    
                model = str(row[model_col - 1] if row[model_col - 1] else "")
                if not model:
                    continue
                    
                if model not in box_map:
                    errors.append(f"Row {row_idx}: Model '{model}' not found in store")
                    continue
                
                box_idx = box_map[model]
                box = current_data["boxes"][box_idx]
                
                # Update itemized prices
                if "itemized-prices" not in box:
                    box["itemized-prices"] = {}
                
                ip = box["itemized-prices"]
                
                # Map header names to price fields
                field_map = {
                    "Box Price": "box-price",
                    "Standard Materials": "standard-materials",
                    "Standard Services": "standard-services",
                    "Fragile Materials": "fragile-materials",
                    "Fragile Services": "fragile-services",
                    "Custom Materials": "custom-materials",
                    "Custom Services": "custom-services"
                }
                
                for header_name, field_name in field_map.items():
                    if header_name in headers:
                        val = row[headers[header_name] - 1]
                        if val is not None:
                            ip[field_name] = float(val)
                
                updated_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        # Close workbook
        wb.close()
        
        # Save updated data
        if updated_count > 0:
            save_yaml_func(store_id, current_data)
        
        # Clean up temp file
        os.unlink(temp_file.name)
        
        return {
            "message": f"Successfully updated {updated_count} box prices",
            "updated": updated_count,
            "errors": errors
        }
        
    except Exception as e:
        # Clean up temp file on error
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


async def analyze_excel_structure(file: UploadFile) -> dict:
    """Analyze Excel file structure for import preview"""
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    # Check file size (512KB limit for production)
    MAX_FILE_SIZE = 512 * 1024  # 512KB
    
    # Stream file to temp location to avoid loading into memory
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        # Stream file content to temp file in chunks
        total_size = 0
        chunk_size = 8192  # 8KB chunks
        
        while True:
            chunk = await file.read(chunk_size)
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > MAX_FILE_SIZE:
                temp_file.close()
                os.unlink(temp_file.name)
                raise HTTPException(status_code=400, detail="File size exceeds 512KB limit")
            temp_file.write(chunk)
        
        temp_file.close()
        await file.seek(0)  # Reset file pointer
        
        # Read Excel file info with minimal memory usage
        wb = load_workbook(temp_file.name, read_only=True, data_only=True, keep_vba=False)
        sheets_info = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value else f"Column{cell.column}")
            
            # Collect data and analyze
            data_rows = []
            column_data = {header: [] for header in headers}
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if row_idx > 51:  # Limit to first 50 data rows for sample
                    break
                
                row_data = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        header = headers[col_idx]
                        row_data[header] = value
                        column_data[header].append(value)
                
                if any(v is not None for v in row):  # Skip completely empty rows
                    data_rows.append(row_data)
            
            # Analyze all rows for statistics
            all_column_data = {header: [] for header in headers}
            total_rows = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                total_rows += 1
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        all_column_data[headers[col_idx]].append(value)
            
            # Build sheet info
            sheet_info = {
                "name": sheet_name,
                "rows": total_rows,
                "columns": headers,
                "shape": (total_rows, len(headers)),
                "sample_data": data_rows,
                "data_types": {},
                "null_counts": {},
                "unique_counts": {},
                "column_analysis": {}
            }
            
            # Analyze each column
            for col_name in headers:
                col_values = all_column_data[col_name]
                non_null_values = [v for v in col_values if v is not None]
                
                # Determine data type
                numeric_count = sum(1 for v in non_null_values if isinstance(v, (int, float)))
                string_count = sum(1 for v in non_null_values if isinstance(v, str))
                
                if numeric_count > string_count:
                    dtype = "numeric"
                else:
                    dtype = "string"
                
                sheet_info["data_types"][col_name] = dtype
                sheet_info["null_counts"][col_name] = len(col_values) - len(non_null_values)
                sheet_info["unique_counts"][col_name] = len(set(non_null_values))
                
                col_analysis = {
                    "dtype": dtype,
                    "null_count": sheet_info["null_counts"][col_name],
                    "unique_count": sheet_info["unique_counts"][col_name],
                    "unique_ratio": sheet_info["unique_counts"][col_name] / len(col_values) if col_values else 0
                }
                
                # For numeric columns, add statistics
                if dtype == "numeric" and non_null_values:
                    numeric_values = [v for v in non_null_values if isinstance(v, (int, float))]
                    if numeric_values:
                        col_analysis["min"] = float(min(numeric_values))
                        col_analysis["max"] = float(max(numeric_values))
                        col_analysis["mean"] = float(sum(numeric_values) / len(numeric_values))
                        sorted_values = sorted(numeric_values)
                        mid = len(sorted_values) // 2
                        if len(sorted_values) % 2 == 0:
                            col_analysis["median"] = float((sorted_values[mid-1] + sorted_values[mid]) / 2)
                        else:
                            col_analysis["median"] = float(sorted_values[mid])
                
                # For string columns, add sample unique values
                if dtype == "string":
                    string_values = [str(v) for v in non_null_values if v is not None]
                    unique_values = list(set(string_values))[:20]
                    col_analysis["sample_unique_values"] = unique_values
                    
                    # Check if it looks like a product/box description
                    if any(keyword in col_name.lower() for keyword in ['box', 'product', 'name']):
                        # Extract potential dimensions
                        dimension_pattern = r'(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)'
                        dimension_examples = []
                        
                        for val in string_values[:100]:  # Check first 100 values
                            if re.search(dimension_pattern, val):
                                dimension_examples.append(val)
                                if len(dimension_examples) >= 10:
                                    break
                        
                        if dimension_examples:
                            col_analysis["contains_dimensions"] = True
                            col_analysis["dimension_examples"] = dimension_examples
                
                sheet_info["column_analysis"][col_name] = col_analysis
            
            sheets_info.append(sheet_info)
        
        # Close workbook
        wb.close()
        
        # Clean up temp file
        os.unlink(temp_file.name)
        
        # Create analysis result
        analysis_result = {
            "filename": file.filename,
            "analysis_date": datetime.now().isoformat(),
            "sheets": sheets_info
        }
        
        # Save to JSON file
        output_filename = f"excel_digest_{file.filename.replace('.xlsx', '').replace('.xls', '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), output_filename)
        
        with open(output_path, 'w') as f:
            json.dump(analysis_result, f, indent=2)
        
        return {
            "message": f"Excel file digested successfully",
            "digest_file": output_filename,
            "summary": {
                "filename": file.filename,
                "sheets": len(sheets_info),
                "total_rows": sum(sheet["rows"] for sheet in sheets_info),
                "digest_saved_to": output_filename
            }
        }
        
    except Exception as e:
        # Clean up temp file on error
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        raise HTTPException(status_code=500, detail=f"Error analyzing file: {str(e)}")


def get_dimensions_from_name(name: str) -> Optional[tuple]:
    """Extract dimensions from a box name like '10x10x10 Box' or '17.5x15.125x3.25'"""
    match = re.search(r'(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)', name)
    if match:
        return (float(match.group(1)), float(match.group(2)), float(match.group(3)))
    return None


def categorize_product(product_name: str) -> str:
    """Helper function to categorize product based on name"""
    name_lower = product_name.lower()
    
    # Check if it's a service (handles abbreviations and full spellings)
    if 'pack service' in name_lower or 'pack svc' in name_lower:
        if 'basic' in name_lower:
            return 'basic_service'
        elif 'std' in name_lower or 'standard' in name_lower:
            return 'standard_service'
        elif 'frg' in name_lower or 'fragile' in name_lower:
            return 'fragile_service'
        elif 'cust' in name_lower or 'custom' in name_lower:
            return 'custom_service'
    # Check if it's materials (handles abbreviations and full spellings)
    elif 'pack material' in name_lower or 'pack mat' in name_lower:
        if 'basic' in name_lower:
            return 'basic_materials'
        elif 'std' in name_lower or 'standard' in name_lower:
            return 'standard_materials'
        elif 'frg' in name_lower or 'fragile' in name_lower:
            return 'fragile_materials'
        elif 'cust' in name_lower or 'custom' in name_lower:
            return 'custom_materials'
    # Check if it's a box - more flexible now
    elif not any(x in name_lower for x in ['pack material', 'pack mat', 'pack service', 'pack svc']):
        # If it has dimensions anywhere in the name (e.g., "10x10x48" or "Golf club 10x10x48"), consider it a box
        import re
        if re.search(r'\d+(?:\.\d+)?\s*[xX]\s*\d+(?:\.\d+)?\s*[xX]\s*\d+(?:\.\d+)?', product_name):
            return 'box'
        elif 'box' in name_lower:
            return 'box'
    
    return 'other'


async def analyze_import_for_matching(
    store_id: str,
    file: UploadFile,
    store_data: dict
) -> dict:
    """Analyze Excel file for three-tier matching with store boxes"""
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    # Check file size (512KB limit for production)
    MAX_FILE_SIZE = 512 * 1024  # 512KB
    
    # Stream file to temp location to avoid loading into memory
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        # Stream file content to temp file in chunks
        total_size = 0
        chunk_size = 8192  # 8KB chunks
        
        while True:
            chunk = await file.read(chunk_size)
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > MAX_FILE_SIZE:
                temp_file.close()
                os.unlink(temp_file.name)
                raise HTTPException(status_code=400, detail="File size exceeds 512KB limit")
            temp_file.write(chunk)
        
        temp_file.close()
        await file.seek(0)  # Reset file pointer
        
        # Read Excel file with minimal memory usage
        wb = load_workbook(temp_file.name, read_only=True, keep_vba=False, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value] = col_idx
        
        
        # Read all rows into memory for processing
        rows_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # Skip empty rows
                row_dict = {}
                for header, col_idx in headers.items():
                    if col_idx <= len(row):
                        row_dict[header] = row[col_idx - 1]
                rows_data.append(row_dict)
        
        
        # Group Excel items by dimensions
        excel_dimension_groups = {}
        excel_dimension_groups_exact = {}  # Keep exact dimensions
        
        
        for row_idx, row in enumerate(rows_data):
            product_name = str(row.get('Product name', ''))
            
            # Extract dimensions using regex (handles decimal dimensions and mixed case X)
            match = re.search(r'(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)', product_name)
            if match:
                # Keep both exact and integer versions
                dims_exact = [match.group(i) for i in range(1, 4)]
                dims_int = [int(float(match.group(i))) for i in range(1, 4)]
                dim_str_exact = f"{dims_exact[0]}x{dims_exact[1]}x{dims_exact[2]}"
                dim_str_int = f"{dims_int[0]}x{dims_int[1]}x{dims_int[2]}"
                
                
                if dim_str_int not in excel_dimension_groups:
                    excel_dimension_groups[dim_str_int] = []
                if dim_str_exact not in excel_dimension_groups_exact:
                    excel_dimension_groups_exact[dim_str_exact] = []
                
                # Categorize the product
                category = categorize_product(product_name)
                
                # Extract suffix (everything after the dimensions)
                suffix = product_name[match.end():].strip()
                
                item_data = {
                    'item_id': row.get('Item'),
                    'product_name': product_name,
                    'price': float(row.get('Price', 0) or 0),
                    'category': category,
                    'dimensions_exact': dim_str_exact,
                    'dimensions_int': dim_str_int,
                    'suffix': suffix  # Store suffix for display in probable matches
                }
                
                excel_dimension_groups[dim_str_int].append(item_data)
                excel_dimension_groups_exact[dim_str_exact].append(item_data)
            else:
                # TODO: Handle special items like "Electronics Insert", "Med Electronics Insert"
                # These may need special categorization or matching logic
                # Discuss with store managers about how to handle inserts vs regular items
                pass
        
        
        # Analyze matches
        perfect_matches = []
        probable_matches = []
        manual_required = []
        incomplete_matches = []
        
        
        for box_idx, box in enumerate(store_data['boxes']):
            box_model = box.get('model', '')
            # Keep original dimensions (may include decimals)
            box_dims_original = box['dimensions']
            box_dims_int = [int(float(d)) for d in box['dimensions']]
            box_dim_str = f"{box_dims_original[0]}x{box_dims_original[1]}x{box_dims_original[2]}"
            box_dim_str_int = f"{box_dims_int[0]}x{box_dims_int[1]}x{box_dims_int[2]}"
            
            
            # Check for perfect match using MPOS_mapping
            if 'MPOS_mapping' in box and box['MPOS_mapping']:
                mapping = box['MPOS_mapping']
                
                # Verify all mapped items exist in Excel
                all_found = True
                mapped_items = {}
                missing_items = []
                
                for field, item_id in mapping.items():
                    found = False
                    # Convert item_id to string for comparison
                    item_id_str = str(item_id)
                    for row in rows_data:
                        # Also convert Excel item to string for comparison
                        if str(row.get('Item', '')) == item_id_str:
                            mapped_items[field] = {
                                'item_id': item_id,
                                'product_name': row.get('Product name'),
                                'price': float(row.get('Price', 0) or 0)
                            }
                            found = True
                            break
                    if not found:
                        all_found = False
                        missing_items.append(f"{field}: {item_id}")
                
                
                if all_found and len(mapped_items) >= 9:  # All 9 required fields including basic_materials and basic_service
                    perfect_matches.append({
                        'box': {
                            'model': box_model,
                            'dimensions': box_dim_str,
                            'dimensions_int': box_dim_str_int
                        },
                        'mapped_items': mapped_items,
                        'current_prices': box.get('itemized-prices', {})
                    })
                    continue
            
            # Check for dimension-based match (use integer dimensions for matching)
            if box_dim_str_int in excel_dimension_groups:
                excel_items = excel_dimension_groups[box_dim_str_int]
                
                
                # Analyze completeness for itemized pricing
                categories_found = {}
                for item in excel_items:
                    if item['category'] != 'other':
                        categories_found[item['category']] = item
                
                required_categories = ['box', 'basic_materials', 'basic_service',
                                     'standard_materials', 'standard_service',
                                     'fragile_materials', 'fragile_service',
                                     'custom_materials', 'custom_service']
                
                missing = [cat for cat in required_categories if cat not in categories_found]
                
                
                if len(missing) == 0:
                    probable_matches.append({
                        'box': {
                            'model': box_model,
                            'dimensions': box_dim_str,
                            'dimensions_int': box_dim_str_int
                        },
                        'excel_items': excel_items,
                        'categories_found': categories_found,
                        'is_complete': True
                    })
                else:
                    # Incomplete category set - goes to incomplete section
                    incomplete_matches.append({
                        'box': {
                            'model': box_model,
                            'dimensions': box_dim_str,
                            'dimensions_int': box_dim_str_int
                        },
                        'excel_items': excel_items,
                        'categories_found': categories_found,
                        'missing_categories': missing,
                        'reason': 'incomplete_categories'
                    })
            else:
                # No dimension match found
                manual_required.append({
                    'box': {
                        'model': box_model,
                        'dimensions': box_dim_str,
                        'dimensions_int': box_dim_str_int
                    },
                    'excel_items': [],
                    'reason': 'no_dimension_match'
                })
        
        # Close workbook
        wb.close()
        
        # Clean up temp file
        os.unlink(temp_file.name)
        
        
        # Return analysis results
        return {
            'store_id': store_id,
            'summary': {
                'total_boxes': len(store_data['boxes']),
                'perfect_matches': len(perfect_matches),
                'probable_matches': len(probable_matches),
                'incomplete_matches': len(incomplete_matches),
                'manual_required': len(manual_required),
                'excel_dimensions': len(excel_dimension_groups)
            },
            'perfect_matches': perfect_matches,
            'probable_matches': probable_matches,
            'incomplete_matches': incomplete_matches,
            'manual_required': manual_required,
            'excel_dimension_groups': excel_dimension_groups,
            'excel_dimension_groups_exact': excel_dimension_groups_exact
        }
        
    except Exception as e:
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        raise HTTPException(status_code=500, detail=f"Error analyzing import: {str(e)}")


def apply_import_updates(
    store_data: dict,
    updates: Dict[str, Any],
    save_yaml_func
) -> dict:
    """
    Apply import updates. Expects:
    {
        "perfect_updates": [{"model": "...", "prices": {...}}],
        "approved_mappings": [{"model": "...", "mapping": {...}, "prices": {...}}],
        "manual_mappings": [{"model": "...", "mapping": {...}, "prices": {...}}]
    }
    """
    
    # Track updates
    updated_count = 0
    results = []
    
    # Helper to find box by model
    def find_box_by_model(model):
        for i, box in enumerate(store_data['boxes']):
            if box.get('model', '') == model:
                return i, box
        return None, None
    
    # Apply perfect match updates (prices only)
    for update in updates.get('perfect_updates', []):
        idx, box = find_box_by_model(update['model'])
        if box:
            # Update itemized prices
            if 'itemized-prices' not in box:
                box['itemized-prices'] = {}
            box['itemized-prices'].update(update['prices'])
            updated_count += 1
            results.append({'model': update['model'], 'status': 'updated'})
    
    # Apply approved mappings (save mapping and update prices)
    for update in updates.get('approved_mappings', []):
        idx, box = find_box_by_model(update['model'])
        if box:
            # Save MPOS mapping
            box['MPOS_mapping'] = update['mapping']
            # Update prices
            if 'itemized-prices' not in box:
                box['itemized-prices'] = {}
            box['itemized-prices'].update(update['prices'])
            updated_count += 1
            results.append({'model': update['model'], 'status': 'mapped_and_updated'})
    
    # Apply manual mappings (save mapping and update prices)
    for update in updates.get('manual_mappings', []):
        idx, box = find_box_by_model(update['model'])
        if box:
            # Save MPOS mapping
            box['MPOS_mapping'] = update['mapping']
            # Update prices
            if 'itemized-prices' not in box:
                box['itemized-prices'] = {}
            box['itemized-prices'].update(update['prices'])
            updated_count += 1
            results.append({'model': update['model'], 'status': 'manually_mapped'})
    
    # Save updated YAML
    if updated_count > 0:
        save_yaml_func(updates['store_id'], store_data)
    
    return {
        'success': True,
        'updated_count': updated_count,
        'results': results
    }


async def discover_boxes_from_prices(file: UploadFile, store_data: dict) -> dict:
    """
    Analyze price sheet to discover box dimensions and suggest matches from library
    
    Returns:
        dict with:
        - discovered_dimensions: List of unique dimensions found
        - library_matches: Exact matches from box library
        - unmatched_dimensions: Dimensions with no library match
        - already_in_store: Dimensions already in the store
    """
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    # Check file size (512KB limit for production)
    MAX_FILE_SIZE = 512 * 1024  # 512KB
    
    # Stream file to temp location to avoid loading into memory
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        # Stream file content to temp file in chunks
        total_size = 0
        chunk_size = 8192  # 8KB chunks
        
        while True:
            chunk = await file.read(chunk_size)
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > MAX_FILE_SIZE:
                temp_file.close()
                os.unlink(temp_file.name)
                raise HTTPException(status_code=400, detail="File size exceeds 512KB limit")
            temp_file.write(chunk)
        
        temp_file.close()
        await file.seek(0)  # Reset file pointer
        
        # Read Excel file with minimal memory usage
        wb = load_workbook(temp_file.name, read_only=True, keep_vba=False, data_only=True)
        ws = wb.active
        
        # Get headers
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value] = col_idx
        
        # Read all rows into memory (reusing pattern from analyze_import_for_matching)
        rows_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # Skip empty rows
                row_dict = {}
                for header, col_idx in headers.items():
                    if col_idx <= len(row):
                        row_dict[header] = row[col_idx - 1]
                rows_data.append(row_dict)
        
        wb.close()
        
        # Group by dimensions (reusing logic from analyze_import_for_matching)
        discovered_boxes = {}  # dimensions_str -> {count, models, prices}
        
        for row in rows_data:
            product_name = str(row.get('Product name', ''))
            
            # Extract dimensions using same regex as analyze_import_for_matching
            match = re.search(r'(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)', product_name)
            if match:
                # Convert to floats and sort largest to smallest
                dims = [float(match.group(i)) for i in range(1, 4)]
                dims.sort(reverse=True)
                dims_str = "x".join([str(int(d)) if d.is_integer() else str(d) for d in dims])
                
                # Categorize the product
                category = categorize_product(product_name)
                
                # Only process boxes (not services or materials)
                if category == 'box':
                    # Extract price
                    price = float(row.get('Price', 0) or 0)
                    
                    # Store discovery info
                    if dims_str not in discovered_boxes:
                        discovered_boxes[dims_str] = {
                            'dimensions': dims,
                            'count': 0,
                            'models': [],
                            'prices': []
                        }
                    
                    discovered_boxes[dims_str]['count'] += 1
                    discovered_boxes[dims_str]['models'].append(product_name)
                    if price > 0:
                        discovered_boxes[dims_str]['prices'].append(price)
        
        # Now match against box library
        library = BoxLibrary()
        
        # Get existing box dimensions to avoid duplicates
        existing_dimensions = set()
        for box in store_data.get('boxes', []):
            dims = sorted(box['dimensions'], reverse=True)
            dims_str = "x".join([str(d) for d in dims])
            existing_dimensions.add(dims_str)
        
        results = {
            'discovered_dimensions': [],
            'library_matches': [],
            'unmatched_dimensions': [],
            'already_in_store': []
        }
        
        # Process each discovered dimension
        for dims_str, info in discovered_boxes.items():
            dims = info['dimensions']
            
            # Check if already in store
            if dims_str in existing_dimensions:
                results['already_in_store'].append({
                    'dimensions': dims,
                    'dimensions_str': dims_str,
                    'count': info['count'],
                    'models': info['models'],
                    'avg_price': sum(info['prices']) / len(info['prices']) if info['prices'] else None
                })
                continue
            
            # Record as discovered
            discovered = {
                'dimensions': dims,
                'dimensions_str': dims_str,
                'count': info['count'],
                'models': info['models'],
                'avg_price': sum(info['prices']) / len(info['prices']) if info['prices'] else None
            }
            results['discovered_dimensions'].append(discovered)
            
            # Look for ALL boxes with these exact dimensions (may have different alternate depths)
            exact_matches = library.find_all_by_dimensions(dims)
            if exact_matches:
                results['library_matches'].append({
                    'discovered': discovered,
                    'library_boxes': exact_matches  # Multiple boxes with same dims but different alternate depths
                })
            else:
                # No match - needs custom box
                results['unmatched_dimensions'].append(discovered)
        
        # Summary statistics
        results['summary'] = {
            'total_boxes_found': len(discovered_boxes),
            'already_in_store': len(results['already_in_store']),
            'new_dimensions': len(results['discovered_dimensions']),
            'exact_matches': len(results['library_matches']),
            'unmatched': len(results['unmatched_dimensions'])
        }
        
        return results
        
    finally:
        # Clean up temp file
        os.unlink(temp_file.name)